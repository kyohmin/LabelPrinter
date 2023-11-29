from os import path
import sys
import os
import openpyxl
import shutil
import pandas as pd
import time
import datetime

# COPY FILE =====================================
shutil.copy("./N_BASE.xlsx", "TEMP.xlsx")
# shutil.copy("./record.xlsx", "RECORDS.xlsx")

# LOAD FILE =====================================
config_name = 'TEMP.xlsx'

if getattr(sys, 'frozen', False):
    application_path = path.dirname(sys.executable)
elif __file__:
    application_path = path.dirname(__file__)

config_path = path.join(application_path, config_name)

def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()

wb = load_workbook(config_path)

# MAIN LOGIC ====================================

# Read txt files
P_Number_file = open("P_Number.txt", "r", encoding="utf-8")
B_Manager_file = open("B_Manager.txt", "r", encoding="utf-8")
R_Number_file = open("R_Number.txt", "r", encoding="utf-8")
B_Serial_file = open("B_Serial.txt", "r", encoding="utf-8")
Sending_file = open("Sending.txt", "r", encoding="utf-8")
Bringing_file = open("Bringing.txt", "r", encoding="utf-8")
B_Date_file = open("B_Date.txt", "r", encoding="utf-8")
C_Serial_file = open("C_Serial.txt", "r", encoding="utf-8")
Size_file = open("Size.txt", "r", encoding="utf-8")
R_Amount_file = open("R_Amount.txt", "r", encoding="utf-8")

P_Number_li, B_Manager_li, R_Number_li, B_Serial_li = [], [], [], []
Sending_li, Bringing_li, B_Date_li, C_Serial_li, Size_li, R_Amount_li = [], [], [], [], [], []

for i in P_Number_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        P_Number_li.append(i.strip("\n"))

for i in B_Manager_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        B_Manager_li.append(i.strip("\n"))

for i in R_Number_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        R_Number_li.append(i.strip("\n"))

for i in B_Serial_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        B_Serial_li.append(i.strip("\n"))

for i in Sending_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        Sending_li.append(i.strip("\n"))

for i in Bringing_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        Bringing_li.append(i.strip("\n"))

for i in B_Date_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        B_Date_li.append(i.strip("\n"))

for i in C_Serial_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        C_Serial_li.append(i.strip("\n"))

for i in Size_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        Size_li.append(i.strip("\n"))

for i in R_Amount_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        R_Amount_li.append(i.strip("\n"))


numbering = 1
for i in range(len(C_Serial_li)):
    ws = wb["template"]
    ws["B1"] = P_Number_li[0] #프로젝트번호
    ws["B2"] = B_Manager_li[0] #사업부담당
    ws["B3"] = R_Number_li[0] #구매요청번호
    ws["B4"] = B_Serial_li[0] #견적의뢰번호
    ws["B5"] = Sending_li[0] #발주번호
    ws["B6"] = Bringing_li[0] #납품번호
    ws["B10"] = B_Date_li[0] #닙품일자
    ws["B7"] = C_Serial_li[i] #품명
    ws["B8"] = Size_li[i] #규격
    ws["B9"] = R_Amount_li[i] #수량
    wb.save("./PRINTING/"+str(numbering)+".xlsx")
    numbering += 1

numbering = 1
for i in range(len(C_Serial_li)):
    config_name = './PRINTING/'+str(numbering)+'.xlsx'
    config_path = path.join(application_path, config_name)
    os.startfile(config_path,'print')
    time.sleep(0.2)
    numbering += 1
    
wb.close()

for i in range(len(C_Serial_li)+1):
    time.sleep(2)
    if os.path.exists("./PRINTING/"+str(i)+".xlsx"):
        os.remove("./PRINTING/"+str(i)+".xlsx")
    else:
        pass

# wb = load_workbook(config_path)
# ws = wb['Sheet1']
# pointer = 2
# for i in range(len(name_li)):
#     ws["A2"] = project_li[0]
#     ws["B2"] = date_li[0]
    
#     name = "C" + str(pointer)
#     number = "D" + str(pointer)
#     quantity = "E" + str(pointer)

#     ws[name] = name_li[i] # ++
#     ws[number] = number_li[i] # ++
#     ws[quantity] = quantity_li[i] # ++
#     pointer += 1

# ymd, hms = str(datetime.datetime.now()).split()
# hms = hms[:2] + '시-' + hms[3:5] + '분'

# abc = './라벨기록/'+ymd+'_'+hms+'.xlsx'
# wb.save(abc)