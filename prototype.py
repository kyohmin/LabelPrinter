from os import path
import sys
import os
import openpyxl
import shutil
import pandas as pd
import time
import datetime

# COPY FILE =====================================
shutil.copy("./BASE.xlsx", "TEMP.xlsx")
# shutil.copy("./record.xlsx", "RECORDS.xlsx")

# LOAD FILE =====================================
config_name = 'TEMP.xlsx'

if getattr(sys, 'frozen', False):
    application_path = path.dirname(sys.executable)
elif __file__:
    application_path = path.dirname(__file__)

config_path = path.join(application_path, config_name)

def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()

wb = load_workbook(config_path)

# MAIN LOGIC ====================================

# Read txt files
project_file = open("project.txt", "r", encoding="utf-8")
date_file = open("date.txt", "r", encoding="utf-8")
name_file = open("name.txt", "r", encoding="utf-8")
number_file = open("number.txt", "r", encoding="utf-8")
quantity_file = open("quantity.txt", "r", encoding="utf-8")

project_li, date_li, name_li, number_li, quantity_li = [], [], [], [], []

for i in project_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        project_li.append(i.strip("\n"))

for i in date_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        date_li.append(i.strip("\n"))

for i in name_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        name_li.append(i.strip("\n"))

for i in number_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        number_li.append(i.strip("\n"))

for i in quantity_file:
    i = i.strip()
    i = i.strip("\n")
    if i == "":
        pass
    else:
        quantity_li.append(i.strip("\n"))

numbering = 1
for i in range(len(name_li)):
    ws = wb["template"]
    ws["B1"] = project_li[0] #프로젝트
    ws["D4"] = date_li[0] #날짜
    ws["B2"] = name_li[i] #품명
    ws["B3"] = number_li[i] #품번
    ws["B4"] = quantity_li[i] #수량
    wb.save("./PRINTING/"+str(numbering)+".xlsx")
    numbering += 1

numbering = 1
for i in range(len(name_li)):
    config_name = './PRINTING/'+str(numbering)+'.xlsx'
    config_path = path.join(application_path, config_name)
    os.startfile(config_path,'print')
    time.sleep(0.2)
    numbering += 1
    
wb.close()

for i in range(len(number_li)+1):
    time.sleep(2)
    if os.path.exists("./PRINTING/"+str(i)+".xlsx"):
        os.remove("./PRINTING/"+str(i)+".xlsx")
    else:
        pass

# wb = load_workbook(config_path)
# ws = wb['Sheet1']
# pointer = 2
# for i in range(len(name_li)):
#     ws["A2"] = project_li[0]
#     ws["B2"] = date_li[0]
    
#     name = "C" + str(pointer)
#     number = "D" + str(pointer)
#     quantity = "E" + str(pointer)

#     ws[name] = name_li[i] # ++
#     ws[number] = number_li[i] # ++
#     ws[quantity] = quantity_li[i] # ++
#     pointer += 1

# ymd, hms = str(datetime.datetime.now()).split()
# hms = hms[:2] + '시-' + hms[3:5] + '분'

# abc = './라벨기록/'+ymd+'_'+hms+'.xlsx'
# wb.save(abc)